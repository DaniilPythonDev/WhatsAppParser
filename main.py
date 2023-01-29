import datetime
import sys
import os
import time
from threading import Thread

from keyboard import add_hotkey
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.common.exceptions import NoSuchWindowException
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains, ScrollOrigin
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from win10toast import ToastNotifier


class Pars:

    def __init__(self, pid):
        self.protsessid = pid
        self.start_pars = False
        self.driver = None
        self.temp_elem = None
        self.close = False
        self.list_contacts = list()
        self.group = None
        self.main_block_chats_list = None
        self.timeout = 1.5
        self.wait = None
        self.count_user = None

    def keyboard_heandler(self):
        try:
            self.run_browser_chrome()
            add_hotkey('ctrl + y', lambda: self.start_script())
            add_hotkey('esc', lambda: self.close_script())
            while True:
                time.sleep(1)
                if self.close:
                    break
                self.driver.get_issue_message()
            # sys.exit(0)
        except NoSuchWindowException:
            os.kill(self.protsessid, 9)
            # sys.exit(0)
        except Exception as ex:
            _ = ex
            # print(f'def keyboard_heandler: {ex}')

    def run_browser_chrome(self):
        try:
            self.driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
            self.driver.get(url='https://web.whatsapp.com/')
            self.driver.implicitly_wait(5)
            self.wait = WebDriverWait(driver=self.driver, timeout=1.5)
        except Exception as _ex:
            _ = _ex
            # print(f'def run_browser_chrome: {_ex}')

    def start_script(self):
        try:
            self.set_element_groupp()
            if self.chat_submenu():
                is_groupp = self.is_groupp()
                if type(is_groupp) == tuple:
                    count = is_groupp[0]
                    self.count_user = count
                    if self.scroll_element(element=self.temp_elem):
                        if count > 10:
                            open_list = self.open_contact_list()
                            time.sleep(1)
                            if open_list and type(open_list) == bool:
                                self.list_contacts = list()
                                self.list_contacts = self.get_contacts_list(count=count)
                                self.list_contacts = list(set(self.list_contacts))
                                thread = Thread(target=WraiteToFile().write_to_sheet,
                                                args=(self.group, self.list_contacts))
                                thread.start()
                                thread.join()
                                self.notification(title='Успех', mess='Сбор данных завершен!')

                        else:
                            self.list_contacts = list()
                            self.get_contakts(all_contacts=self.list_contacts, little=True)
                            self.list_contacts = list(set(self.list_contacts))
                            thread = Thread(target=WraiteToFile().write_to_sheet,
                                            args=(self.group, self.list_contacts))
                            thread.start()
                            thread.join()
                            self.notification(title='Успех', mess='Сбор данных завершен!')
                else:
                    ...
        except Exception as _ex:
            _ = _ex
            # print(f'def start_script: {_ex}')

    def chat_submenu(self):
        try:
            submenu = self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[class="_24-Ff"]')))
            if submenu:
                submenu.click()
                return True
        except Exception as _ex:
            _ = _ex
            # print(f'def chat_submenu: {_ex}')
            return False

    def is_groupp(self):
        try:
            is_gruppe = self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR,
                                                                    '[class="a4ywakfo ma4rpf0l qfejxiq4"]')))
            if is_gruppe.text.strip().startswith('Группа'):
                self.temp_elem = is_gruppe
                count = str(is_gruppe.text.strip().split()[-2])
                if count.isdigit():
                    return int(count) - 1, True
        except Exception as _ex:
            self.notification(title='Внимание', mess='Войдите в группу!')
            _ = _ex
            # print(f'def is_group: {_ex}')
            return False

    def scroll_element(self, element, y=10000, scroll_to=False):
        try:
            actions = ActionChains(self.driver)
            if not scroll_to:
                scroll_origin = ScrollOrigin.from_element(element)
                actions.scroll_from_origin(scroll_origin, 0, y).perform()
            else:
                actions.scroll_to_element(element=element).perform()

            return True
        except Exception as _ex:
            _ = _ex
            # print(f'def scroll_to_element: {_ex}')
            return False

    def open_contact_list(self):
        try:
            block_list_contackts = self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR,
                                                                               '[class="i5tg98hk f9ovudaz przvwfww ddw6'
                                                                               's8x9 shdiholb phqmzxqs pm5hny62 ajgl1lb'
                                                                               'b thr4l2wc cc8mgx9x eta5aym1 d9802myq e'
                                                                               '4xiuwjv q1n4p668 ln8gz9je p357zi0d gndf'
                                                                               'cl4n os03hap6"]')))
            if block_list_contackts and str(block_list_contackts.text).strip().startswith('Просмотреть') or \
                    str(block_list_contackts.text).strip().lower().startswith('ещё'):
                block_list_contackts.click()
                return True
            else:
                return False
        except Exception as _ex:
            _ = _ex
            # print(f'def open_contact_list: {_ex}')
            return False

    def get_contakts(self, all_contacts: list, little: bool = False):
        try:
            for u in range(len(self.blocks_chat_user_list(little=little))):
                if len(all_contacts) == self.count_user:
                    break
                get_values = self.blocks_chat_user_list(little=little)
                text_value = str(get_values[u].text.strip())
                if text_value.translate({ord(i): None for i in ('+', ' ', '-')}).isdigit():
                    if not all_contacts.count((None, text_value)):
                        # print(text_value)
                        all_contacts.append((None, text_value))
                else:
                    if not text_value.lower().startswith('вы'):
                        self.scroll_element(scroll_to=True, element=get_values[u])
                        time.sleep(0.7)
                        get_values[u].click()
                        time.sleep(1)
                        if not little and self.count_user >= 20:
                            menu = self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[class="_2sDI2"]')))
                            click_kategori_send_mess = menu.find_elements(by=By.TAG_NAME,
                                                                          value='li')
                            for e in click_kategori_send_mess:
                                if 'Написать контакту' in e.text.strip():
                                    e.click()
                                    time.sleep(1)
                        data = self.open_and_get_user_data()
                        if not all_contacts.count(data):
                            # print(data[1])
                            all_contacts.append(data)
                        if self.click_main_block_list_chat():
                            self.chat_submenu()
                            self.temp_elem = self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR,

                                                                                         '[class="p357zi0d ktfrpxia nu7'
                                                                                         'pwgvd fhf7t426 f8m0rgwh gndfc'
                                                                                         'l4n"]')))
                            self.scroll_element(element=self.temp_elem, y=5000)
                            if not little:
                                self.open_contact_list()
                                time.sleep(1)
            return {'list': all_contacts}
        except Exception as _ex:
            _ = _ex
            # print(f'def get_contakts: {_ex}')
            return {}

    def get_contacts_list(self, count):
        all_contacts_list = list()
        try:
            ink = 1
            while len(all_contacts_list) < count:
                contacts_list = self.get_contakts(all_contacts=all_contacts_list)
                if contacts_list.get('list'):
                    all_contacts_list = contacts_list.get('list')
                if len(contacts_list.get('list')) < count:
                    list_priwats_chats = self.blocks_chat_user_list()
                    self.scroll_element(element=list_priwats_chats[0], y=1000 * ink)
                    ink += 1
                    time.sleep(1)
                    continue
                else:
                    break
            return all_contacts_list
        except Exception as _ex:
            _ = _ex
            # print(f'def get_contacts_list: {_ex}')

    def close_script(self):
        self.driver.close()
        self.driver.quit()
        # self.close = True
        os.kill(self.protsessid, 9)

    def open_and_get_user_data(self):
        try:
            if self.chat_submenu():
                time.sleep(1)
                try:
                    data_block = self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR,
                                                                             '[class="gsqs0kct oauresqk efgp0a3n h3bz2v'
                                                                             'by g0rxnol2 tvf2evcx oq44ahr5 lb5m6g5c br'
                                                                             'ac1wpa lkjmyc96 b8cdf3jl bcymb0na myel2vf'
                                                                             'b e8k79tju"]')))
                    name = data_block.find_element(by=By.CSS_SELECTOR,
                                                   value='[class="zu5D5 dd2Ow qfejxiq4"]')
                    tel = data_block.find_element(by=By.CSS_SELECTOR,
                                                  value='[class="a4ywakfo qt60bha0"]')

                    if str(tel.text.strip()).translate({ord(i): None for i in (' ', '-', '+')}).isdigit() and name:
                        return name.text.strip(), tel.text.strip()
                    elif str(tel.text.strip()).translate({ord(i): None for i in (' ', '-', '+')}).isdigit():
                        return None, tel.text.strip()
                except Exception as _:
                    _ = _
                    data_block = self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR,
                                                                             '[class="tt8xd2xn jnwc1y2a ngycyvoj svoq16'
                                                                             'ka"]')))
                    name = data_block.find_element(by=By.CSS_SELECTOR,
                                                   value='[class="iqrewfee sy6s5v3r tt8xd2xn jnwc1y2a or9x5nie svoq16k'
                                                         'a"]').text.strip().replace('~', '')
                    self.scroll_element(element=data_block)
                    tel = self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR,
                                                                      '[class="_2vQWV p357zi0d gndfcl4n k45dudtp f9ovud'
                                                                      'az cc8mgx9x"]'))).text.strip()
                    return name, tel
        except Exception as _ex:
            _ = _ex
            # print(f'def open_and_get_user_data: {_ex}')
            return None, None

    def set_element_groupp(self):
        try:
            self.main_block_chats_list = self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR,
                                                                                     '[class="g0rxnol2 _3fGK2"]')))
            block_groupp = self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR,
                                                                       '[aria-selected="true"]')))
            self.group = block_groupp.find_element(by=By.CSS_SELECTOR,
                                                   value='[class="_21S-L"]').text.strip()
            return True
        except Exception as _ex:
            _ = _ex
            # print(f'def set_element_groupp: {_ex}')
            return False

    def click_main_block_list_chat(self):
        try:
            try:
                elem = self.wait.until(
                    EC.element_to_be_clickable((By.CSS_SELECTOR, '[class="p357zi0d ktfrpxia nu7pwgvd fhf'
                                                                 '7t426 f8m0rgwh gndfcl4n"]')))
                self.scroll_element(element=elem, y=5000)
                time.sleep(1)
            except Exception as _:
                _ = _

            self.main_block_chats_list = self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR,
                                                                                     '[class="tt8xd2xn dl6j7rsh mpdn4n'
                                                                                     'r2 avk8rzj1"]')))
            chats_list = self.main_block_chats_list.find_elements(by=By.CSS_SELECTOR,
                                                                  value='[class="_8nE1Y"]')
            for chat in chats_list:
                text = chat.find_element(by=By.CSS_SELECTOR,
                                         value='[class="_21S-L"]').text.strip()
                if self.group == text:
                    chat.click()
                    return True
            return False
        except Exception as _ex:
            _ = _ex
            # print(f'def click_main_block_list_chat: {_ex}')
            return False

    def blocks_chat_user_list(self, little: bool = False):
        try:
            if little:
                main_block_contackts_list = self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR,
                                                                                        '[class="tt8xd2xn dl6j7rsh mpd'
                                                                                        'n4nr2 avk8rzj1"]')))
            else:
                main_block_contackts_list = self.wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR,
                                                                                        '[class="g0rxnol2 g0rxnol2 thgh'
                                                                                        'mljt p357zi0d rjo8vgbg ggj6'
                                                                                        'brxn f8m0rgwh gfz4du6o ag5g9lr'
                                                                                        'v bs7a17vp ov67bkzj"]')))

            get_values = main_block_contackts_list.find_elements(by=By.CSS_SELECTOR, value='[class="_21S-L"]')
            return get_values
        except Exception as _ex:
            _ = _ex
            # print(f'def blocks_chat_user_list: {_ex}')
            return []

    @staticmethod
    def notification(title, mess, duration=4):
        try:
            toaster = ToastNotifier()
            toaster.show_toast(title=title, msg=mess, duration=duration, icon_path='data/ic.ico')
        except Exception as _:
            _ = _


class WraiteToFile:
    def __init__(self):
        self.text_date = f'{str(datetime.datetime.now().date()).replace("-", "_")}'
        self.book = Workbook()
        self.book.remove(self.book.active)

    def write_to_sheet(self, title: str, list_value: list):
        try:
            self.book.create_sheet(title=title)
            sh = self.book[f'{title}']
            sh[f'B1'], sh[f'A1'] = 'Имена:', 'Телефоны:'
            sh[f'B1'].fill = PatternFill("solid", fgColor='FFFF00')
            sh[f'A1'].fill = PatternFill("solid", fgColor='FFFF33')
            sh.column_dimensions['A'].width = 20
            sh.column_dimensions['B'].width = 25
            for i, value in enumerate(list_value):
                value = list(value)
                if not value[0]:
                    value[0] = ''
                if not value[1]:
                    value[1] = ''
                sh[f'B{i + 2}'], sh[f'A{i + 2}'] = value
            path = '../output/'
            if not os.path.exists(path):
                os.mkdir(path)
            file_name = f'{self.text_date}_{title.translate({ord(i): "_" for i in ("-", " ", "/", ".", ",")})}.xlsx'
            self.book.save(filename=f'{path}{file_name}')
        except Exception as _:
            # print(_)
            _ = _


if __name__ == '__main__':
    p = os.getpid()
    Pars(pid=p).keyboard_heandler()
    # WraiteToFile().write_to_sheet(title='dsa', list_value=[(1, 1), (1, 2), (1, 3)])
